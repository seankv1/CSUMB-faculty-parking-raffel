from tkinter import *
from openpyxl import workbook, load_workbook
import pandas as pd
from PIL import Image, ImageTk
import random
root = Tk()
root.title('CSUMB App')
#root.geometry("600x1000")
#root.config(bg="#008B8B")
root.resizable(0, 0)
frame=Frame(root)
frame.pack()
canvas=Canvas(frame,bg="#008B8B", width=600, height=700)
canvas.pack()
img = Image.open('logo4.png') # Allows us to open CSUMB logo image #
img1 = img.convert("RGBA") # Converts image to an RGBA format #
img2 = img1.getdata() # Allows to pull RGBA data from the image #
newData = []
for item in img2:
   if item[0] > 155 and item[1] > 155 and item[2] > 155: # Pulls the white RGB value from the CSUMB logo #
       newData.append((255, 255, 255, 0)) # Then replaces the white with transparency/red RGB value #
   else:
       newData.append(item)
img1.putdata(newData) # Applies changes to the image
img1.save("logo5.png") # Saves the new image with the updated data #
img = Image.open("logo5.png")
img1 = img.resize((100, 100), Image.ANTIALIAS)
img2 = ImageTk.PhotoImage(img1)
#label1=Label(root, image = img2).place(x=200, y =10)
canvas.create_image(75,75, image=img2)
def calendar():
    wb = load_workbook('examexcel.xlsx')
    ws = wb.active
    range = ws['B2':'B21']
    apple = []
    for cell in range:
        for x in cell:
            apple.append(x.value)
            apple.append('\r')
    applestring="".join(apple)
    print(applestring)
    label1.config(text=applestring)
button1 = Button(text="Calender",bg="#76EE00", command=calendar)
button1.place(x=150, y=50)
label1 = Label(bg="#76EE00", height=30, width=50)
label1.place(x=150, y=100)
def buildings():
   wb = load_workbook('examexcel.xlsx')
   ws = wb.active
   range = ws['A2':'A8']
   apple = []
   for cell in range:
       for x in cell:
           apple.append(x.value)
           apple.append('\r')
   applestring = "".join(apple)
   print(applestring)
   label1.config(text=applestring)
button2 = Button(text="Building",bg="#76EE00" ,command=buildings)
button2.place(x=260, y = 50)
label1 = Label(bg="#76EE00", height=30, width=50)
label1.place(x=150, y=100)
def faculty():
   wb = load_workbook('examexcel.xlsx')
   ws = wb.active
   range = ws['C2':'C13']
   apple = []
   for cell in range:
       for x in cell:
           apple.append(x.value)
           apple.append('\r')
   applestring = "".join(apple)
   print(applestring)
   label1.config(text=applestring)
button3 = Button(text="Faculty",bg="#76EE00" ,command=faculty)
button3.place(x=373, y = 50)
def lucky_faculty():
   wb = load_workbook('examexcel.xlsx')
   ws = wb.active
   range = ws['C2':'C13']
   apple = []
   for cell in range:
       for x in cell:
           apple.append(x.value)
   print(apple)
   computer_action = random.choice(apple)
   print(computer_action)
   label1.config(text=computer_action + ' has won the free parking raffel for this month!!')
button4 = Button(text="Lucky Faculty",bg="#76EE00" ,command=lucky_faculty)
button4.place(x=473, y = 50)
root.mainloop()














